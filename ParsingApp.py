# -*- coding: utf-8 -*-
# PYTHON 3.6, 3.7, 3.9

# pyinstaller --icon=pngegg.ico --onefile main.py

# 개선할 점
# 1. utf-8 변환 에러 수정
# 2. 페이지 나누기 미리보기

import io, os, time, sys, re
import warnings
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import io, os, time, sys

warnings.simplefilter("ignore")



def ConvertFile(pathScriptDir, fileList) :
    for fileName in fileList:
        fileDir = pathScriptDir + '\\' + fileName
        try:
            file = io.open(fileDir, mode='r').read()
        except UnicodeError:
            continue
        io.open(fileDir, mode='w', encoding="utf-8").write(file)



def InputInformation():
    menu = input("* 파싱 결과 양식 선택\n  하나의 시트로 생성:1     (sample 시트 필요)하나의 시트로 생성:2     (sample 시트 필요)여러 시트로 생성:3\n  => ")

    sampleReportFile = ""
    if menu == '2' or menu == '3':
        print(menu)
        sampleReportFile = input("* 상세결과 파일 드래그 또는 입력 (sample 시트 필요)\n  Ex) C:\\Users\hyebeen\\Desktop\\01_OOO_시스템 취약점 진단 결과_Unix_v0.1.xlsx \n  => ")
        sampleReportFile = sampleReportFile.strip('"')

    pathScriptDir = input("* 스크립트 결과 폴더 드래그 또는 입력\n  Ex) C:\\Users\hyebeen\\Desktop\\directory \n  => ")
    try :
        fileList = os.listdir(pathScriptDir)
    except FileNotFoundError :
        print("* 잘못된 경로")
        time.sleep(2)
        sys.exit()

    startLine = input("* \"시작 패턴\"을 입력\n  (default) ##### START ##### \n  =>  ")
    if startLine == "" : startLine = "##### START #####"

    endLine = input("* \"마지막 패턴\"을 입력\n  (default) ##### END ##### \n  =>  ")
    if endLine == "" : endLine = "##### END #####"

    return pathScriptDir, sampleReportFile, fileList, startLine, endLine, menu



def RemoveFile() :
    if os.path.isfile("(파싱결과) 시스템 취약점 진단 결과.xlsx"):
        removeFlag = input("* 기존 파일 \"(파싱결과) 시스템 취약점 진단 결과.xlsx\"이 존재합니다.\n* 계속 진행하려면 삭제해야 합니다. 삭제하시겠습니까?\n  예:y, 아니오:n\n =>  ")
        if removeFlag == "y" :
            try:
                os.remove("(파싱결과) 시스템 취약점 진단 결과.xlsx")
            except PermissionError:
                print("* \"(파싱결과) 시스템 취약점 진단 결과.xlsx\"파일을 닫은 후 재실행해주세요.\n")
                time.sleep(2)
                sys.exit()
            print("* 기존 파일 \"(파싱결과) 시스템 취약점 진단 결과.xlsx\"을 삭제하였습니다.")
        if removeFlag == "n":
            print("기존 파일 \"(파싱결과) 시스템 취약점 진단 결과.xlsx\" 삭제 후 실행해야 합니다.\n")
            time.sleep(2)
            sys.exit()



def SetStyle(ws, rowIndex):
    ScolIndex = 11
    SrowIndex = rowIndex

    # 행 높이 설정
    for row in range(1, rowIndex + 1):
        ws.row_dimensions[row].height = 18

    # 열 너비 설정
    ws.column_dimensions['J'].width = 47

    # 표 제목 글자
    for r in range(1, ScolIndex+1):
        ws.cell(1, r).font = Font(bold=True, size=9, color='FFFFFF')
        ws.cell(1, r).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(1, r).fill = PatternFill(start_color='404040', end_color='404040', fill_type='solid')

    # 표 내용 글자
    for i in range(2, SrowIndex+1):
        for s in range(1, ScolIndex+1):
            ws.cell(i, s).font = Font(size=9, color='000000')


    # 테두리 설정
    # 표 내용 테두리
    for i in range(2, SrowIndex+1):
        for s in range(1, ScolIndex+1):
            ws.cell(i, s).border = Border(left =Side(style='thin', color='D9D9D9'),
                                            right = Side(style='thin', color='D9D9D9'),
                                            top=Side(style='thin', color='D9D9D9'),
                                            bottom=Side(style='thin', color='D9D9D9'))
    # 표 제목 테두리
    for s in range(1, ScolIndex+1):
        ws.cell(1, s).border = Border(left =Side(style='thin', color='D9D9D9'),
                                        right = Side(style='thin', color='D9D9D9'),
                                        top=Side(style='medium', color='002060'),
                                        bottom=Side(style='thin', color='002060'))



def firstMenu(pathScriptDir, fileList, startLine, endLine) :
    # 엑셀 시트 및 기본 틀 생성
    wb = Workbook()
    ws = wb.active

    ws.sheet_view.showGridLines = False

    ws.title = "파싱 결과"

    ws.cell(1, 1, "세부 항목")
    ws.cell(1, 2, "CODE")
    ws.cell(1, 3, "점검항목")
    ws.cell(1, 4, "위험도")
    ws.cell(1, 5, "Hostname")
    ws.cell(1, 6, "IP")
    ws.cell(1, 7, "분류")
    ws.cell(1, 8, "진단결과")
    ws.cell(1, 9, "이행결과")
    ws.cell(1, 10, "현재설정")
    ws.cell(1, 11, "비고")

    readFlag = "False"
    write_txt = ""
    rowIndex = 1
    colIndex = 10
    resultFile = "상세결과.xlsx"

    # 파일 하나씩 열어서 파싱
    for fileName in fileList:
        fileDir = pathScriptDir + '\\' + fileName
        txt_f = open(fileDir, encoding="utf-8", mode="r")

        while True:
            txt = txt_f.readline()
            if startLine in txt:
                readFlag = "True"
                rowIndex += 1
                continue
            if endLine in txt:
                readFlag = "False"
                ws.cell(rowIndex, 5, re.search('.+_[0-9]+.[0-9]+.[0-9]+.[0-9]+_(.+).log', fileName).group(1))
                ws.cell(rowIndex, 6, re.search('.+_([0-9]+.[0-9]+.[0-9]+.[0-9]+)_.+.log', fileName).group(1))
                ws.cell(rowIndex, 7, re.search('(.+)_[0-9]+.[0-9]+.[0-9]+.[0-9]+_(.+).log', fileName).group(1))
                ws.cell(rowIndex, colIndex, write_txt)
                ws.cell(rowIndex, colIndex).alignment = Alignment(wrap_text=True)
                write_txt = txt
            if readFlag == "True":
                write_txt += txt
            if not txt: break

        txt_f.close()

    # 스타일 설정
    SetStyle(ws, rowIndex)

    wb.save(filename="(파싱결과) 시스템 취약점 진단 결과.xlsx")

    # 파일 저장
    wb.save(resultFile)



def secondMenu(pathScriptDir, sampleReportFile, fileList, startLine, endLine) :

    wb = load_workbook(sampleReportFile)
    copy_ws = wb.copy_worksheet(wb["sample"])
    copy_ws.title = "파싱 결과"
    copy_ws.sheet_view.showGridLines = False

    readFlag = "False"

    #sample 시트 수정해서 틀 만들기
    copy_ws.unmerge_cells('A1:A3')
    copy_ws.unmerge_cells('E1:F1')
    copy_ws.unmerge_cells('E2:F2')
    copy_ws.unmerge_cells('E3:F3')

    copy_ws.delete_rows(1, 4)
    copy_ws.row_dimensions[4].height = 18
    copy_ws.column_dimensions["I"].hidden=False

    copy_ws.cell(1, 9, "Hostname")
    copy_ws.cell(1, 10, "IP")
    copy_ws.cell(1, 11, "분류")


    rowIndex = 1
    colIndex = 7

    for fileName in fileList:
        fileDir = pathScriptDir + '\\' + fileName
        txt_f = open(fileDir, encoding="utf-8", mode="r")

        while True:
            txt = txt_f.readline()
            if startLine in txt :
                write_txt =""
                readFlag = "True"
                rowIndex += 1
                continue
            if endLine in txt :
                readFlag = "False"
                copy_ws.cell(rowIndex, 9, re.search('.+_[0-9]+.[0-9]+.[0-9]+.[0-9]+_(.+).log', fileName).group(1))
                copy_ws.cell(rowIndex, 10, re.search('.+_([0-9]+.[0-9]+.[0-9]+.[0-9]+)_.+.log', fileName).group(1))
                copy_ws.cell(rowIndex, 11, re.search('(.+)_[0-9]+.[0-9]+.[0-9]+.[0-9]+_(.+).log', fileName).group(1))
                copy_ws.cell(rowIndex, colIndex, write_txt)
                copy_ws.cell(rowIndex, colIndex).alignment = Alignment(wrap_text=True)
            if readFlag == "True":
                write_txt += txt
            if not txt: break

        txt_f.close()

    # 행 높이 설정
    for row in range(1, rowIndex + 1):
        copy_ws.row_dimensions[row].height = 18

    #스타일설정
    ScolIndex = 11
    SrowIndex = rowIndex

    #표 제목 글자
    for i in range(1, ScolIndex+1):
        copy_ws.cell(1, i).font = Font(bold=True, size=9, color='FFFFFF')
        copy_ws.cell(1, i).alignment = Alignment(horizontal='center', vertical='center')
        copy_ws.cell(1, i).fill = PatternFill(start_color='404040', end_color='404040', fill_type='solid')

    # 표 내용 글자
    for i in range(2, SrowIndex+1):
        for s in range(1, ScolIndex+1):
            copy_ws.cell(i, s).font = Font(size=9, color='000000')


    # 테두리 설정
    # 표 내용 테두리
    for i in range(1, SrowIndex+1):
        for s in range(1, ScolIndex+1):
            copy_ws.cell(i, s).border = Border(left =Side(style='thin', color='D9D9D9'),
                                            right = Side(style='thin', color='D9D9D9'),
                                            top=Side(style='thin', color='D9D9D9'),
                                            bottom=Side(style='thin', color='D9D9D9'))

    # 파일 저장
    wb.save(filename="(파싱결과) 시스템 취약점 진단 결과.xlsx")



def thirdMenu(pathScriptDir, sampleReportFile, fileList, startLine, endLine) :
    wb = load_workbook(sampleReportFile)
    readFlag = "False"

    for fileName in fileList:
        fileDir = pathScriptDir + '\\' + fileName
        txt_f = open(fileDir, encoding="utf-8", mode="r")

        copy_ws = wb.copy_worksheet(wb["sample"])
        copy_ws.title = re.search('.*_([a-zA-Z0-9]+).log', fileName).group(1)
        copy_ws.sheet_view.showGridLines = False

        rowIndex = 5
        colIndex = 7

        while True:
            txt = txt_f.readline()
            if startLine in txt:
                write_txt = ""
                readFlag = "True"
                rowIndex += 1
                continue
            if endLine in txt:
                readFlag = "False"
                copy_ws.cell(rowIndex, colIndex, write_txt)
                copy_ws.cell(rowIndex, colIndex).alignment = Alignment(wrap_text=True)
            if readFlag == "True":
                write_txt += txt
            if not txt: break

    # 파일 저장
    wb.save(filename="(파싱결과) 시스템 취약점 진단 결과.xlsx")



def main():
    # 기존 파일 있으면 삭제
    RemoveFile()

    # 정보 입력
    pathScriptDir, sampleReportFile, fileList, startLine, endLine, menu = InputInformation()

    # utf-8로 변환
    ConvertFile(pathScriptDir, fileList)

    # 파싱 실행
    if menu == '1':
        firstMenu(pathScriptDir, fileList, startLine, endLine)
    elif menu == '2':
        secondMenu(pathScriptDir, sampleReportFile, fileList, startLine, endLine)
    elif menu == '3':
        thirdMenu(pathScriptDir, sampleReportFile, fileList, startLine, endLine)

    # 종료
    print("*** [완료] \"파싱 프로그램\"이 있는 경로에 \"(파싱결과) 시스템 취약점 진단 결과.xlsx\" 결과 파일 생성 ***")
    time.sleep(5)



if __name__ == '__main__':
    main()
